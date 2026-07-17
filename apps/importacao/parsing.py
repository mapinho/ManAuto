"""Utilitarios de leitura e parsing de planilhas de importacao (templates Vector).

Regras gerais (skill agrovector-dados): primeira linha = cabecalho; aceitar
CSV UTF-8 com/sem BOM; aceitar decimal com ponto OU virgula (dados reais
chegam em ambos os formatos, ver `template_historico_os.xlsx`); corrigir
mojibake conhecido (`JoÃ£o` -> `João`) antes de validar.
"""

from __future__ import annotations

import csv
from collections.abc import Iterator
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl


@dataclass(frozen=True)
class ErroLinha:
    linha: int
    mensagem: str


@dataclass
class ResultadoImportacao:
    total_linhas: int = 0
    importados: int = 0
    erros: list[ErroLinha] = field(default_factory=list)

    @property
    def teve_erros(self) -> bool:
        return bool(self.erros)

    def registrar_erro(self, linha: int, mensagem: str) -> None:
        self.erros.append(ErroLinha(linha=linha, mensagem=mensagem))


def normalizar_mojibake(texto: str | None) -> str | None:
    """Corrige mojibake de UTF-8 lido como CP1252 (`Ã£` -> `ã`).

    Dados reais dos templates Vector chegam com essa corrupcao (bytes UTF-8
    decodificados errado e re-salvos) — ver skill agrovector-dados. Usa
    CP1252 (nao Latin-1/ISO-8859-1): e o codepage padrao do Windows/Excel,
    origem real da corrupcao, e cobre caracteres tipograficos (aspas curvas,
    travessao) na faixa 0x80-0x9F que o Latin-1 nao representa — um arquivo
    Vector real tinha "Ó" corrompido usando exatamente essa faixa, e a
    tentativa com Latin-1 falhava silenciosamente (round-trip nao fechava) e
    deixava o mojibake sem corrigir. So corrige quando o round-trip fecha
    (texto realmente veio de UTF-8 mal decodificado); caso contrario devolve
    o texto original sem alterar.
    """
    if not texto:
        return texto
    try:
        corrigido = texto.encode("cp1252").decode("utf-8")
    except (UnicodeDecodeError, UnicodeEncodeError):
        return texto
    return corrigido


def _normalizar_valor(valor: Any) -> Any:
    if isinstance(valor, str):
        valor = normalizar_mojibake(valor.strip())
        return valor if valor != "" else None
    return valor


def ler_linhas(caminho: str | Path) -> Iterator[tuple[int, dict[str, Any]]]:
    """Le uma planilha (XLSX ou CSV), primeira linha = cabecalho.

    Gera tuplas (numero_da_linha, {coluna: valor}) a partir da segunda linha
    (numero_da_linha=2 para a primeira linha de dados, igual ao Excel).
    """
    caminho = Path(caminho)
    if caminho.suffix.lower() == ".csv":
        yield from _ler_csv(caminho)
    else:
        yield from _ler_xlsx(caminho)


def _ler_xlsx(caminho: Path) -> Iterator[tuple[int, dict[str, Any]]]:
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        linhas = ws.iter_rows(values_only=True)
        cabecalho = [str(c).strip() if c else "" for c in next(linhas)]
        for i, linha in enumerate(linhas, start=2):
            if all(v is None for v in linha):
                continue
            registro = {
                cabecalho[j]: _normalizar_valor(v)
                for j, v in enumerate(linha)
                if j < len(cabecalho) and cabecalho[j]
            }
            yield i, registro
    finally:
        wb.close()


def _ler_csv(caminho: Path) -> Iterator[tuple[int, dict[str, Any]]]:
    with caminho.open(encoding="utf-8-sig", newline="") as f:
        leitor = csv.reader(f)
        cabecalho = [c.strip() for c in next(leitor)]
        for i, linha in enumerate(leitor, start=2):
            if not any(v.strip() for v in linha):
                continue
            registro = {
                cabecalho[j]: _normalizar_valor(v)
                for j, v in enumerate(linha)
                if j < len(cabecalho) and cabecalho[j]
            }
            yield i, registro


def campo_obrigatorio(registro: dict[str, Any], nome: str) -> Any:
    valor = registro.get(nome)
    if valor in (None, ""):
        raise ValueError(f"campo obrigatório '{nome}' vazio")
    return valor


def parse_decimal(valor: Any, *, padrao: Decimal | None = None) -> Decimal:
    """Aceita int/float/str, com decimal em ponto OU vírgula pt-BR."""
    if valor is None or valor == "":
        if padrao is not None:
            return padrao
        raise ValueError("valor numérico vazio")
    if isinstance(valor, int | float | Decimal):
        return Decimal(str(valor))
    texto = str(valor).strip().replace(" ", "")
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    else:
        texto = texto.replace(",", ".")
    try:
        return Decimal(texto)
    except InvalidOperation as exc:
        raise ValueError(f"valor numérico inválido: {valor!r}") from exc


def parse_int(valor: Any, *, padrao: int | None = None) -> int:
    if valor is None or valor == "":
        if padrao is not None:
            return padrao
        raise ValueError("valor inteiro vazio")
    try:
        return int(parse_decimal(valor))
    except ValueError as exc:
        raise ValueError(f"valor inteiro inválido: {valor!r}") from exc


def parse_bool(valor: Any, *, padrao: bool | None = None) -> bool:
    if valor is None or valor == "":
        if padrao is not None:
            return padrao
        raise ValueError("valor booleano vazio")
    if isinstance(valor, bool):
        return valor
    texto = str(valor).strip().lower()
    if texto in ("true", "sim", "s", "1", "yes"):
        return True
    if texto in ("false", "não", "nao", "n", "0", "no"):
        return False
    raise ValueError(f"valor booleano inválido: {valor!r}")


def parse_data(valor: Any) -> date:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        texto = valor.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(texto, fmt).date()
            except ValueError:
                continue
    raise ValueError(f"data inválida: {valor!r}")
